import pprint
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


bottom_border = Border(bottom=Side(style='thick'))
left_border = Border(left=Side(style='thick'))
right_border = Border(right=Side(style='thin'))
bottom_right_border = Border(right=Side(style='thick'), 
                     bottom=Side(style='thick'))
right_left_border=Border(right=Side(style='thick'), 
                     left=Side(style='thick'))
bottom_top_right_border=Border(right=Side(style='thick'), 
                     bottom=Side(style='thick'),
                     top=Side(style="thick"))
bottom_top_border=Border(bottom=Side(style='thick'),
                     top=Side(style="thick"))



component_dict = {
    'ELE': [
        'battery', 'alternator', 'starter', 'spark plug', 'coil', 'fuse',
        'wiring', 'harness', 'relay', 'sensor', 'actuator', 'motor',
        'diode', 'circuit', 'ecu', 'module', 'light', 'bulb', 'terminal',
        'ground', 'voltage', 'amperage', 'resistance','a/c', 'air conditioning', 'heater',
        'climate control', 'compressor', 'evaporator', 'condenser', 'blower', 'fan', 'wire'
    ],
    'TYRE': [
        'tire', 'tyre', 'wheel', 'rim', 'rubber', 'tread', 'sidewall',
        'pressure', 'valve', 'bead', 'balance', 'rotation', 'puncture',
        'flat', 'blistering', 'graining', 'lock-up', 'marbles'
    ],
    'MECH': [
        'engine', 'transmission', 'gearbox', 'suspension', 'brake',
        'clutch', 'axle', 'driveshaft', 'piston', 'crankshaft',
        'camshaft', 'belt', 'hose', 'radiator', 'pump', 'fluid',
        'oil', 'filter', 'gasket', 'bearing', 'pulley', 'manifold',
        'exhaust', 'catalytic', 'differential',"water", 'cooling', 'steering', 'rack', 'pinion', 'linkage',
        'shock absorber', 'strut', 'spring', 'bushing', 'joint',"air leak","air leaks","gear", "gears", "gearbox", "gear box"
    ],
    'BODY': [
        'chassis', 'frame', 'door', 'hood', 'trunk', 'fender', 'bumper',
        'quarter panel', 'roof', 'pillar', 'windshield', 'glass',
        'mirror', 'handle', 'latch', 'seat', 'carpet', 'dashboard',
        'console', 'trim', 'paint', 'scratch', 'dent', 'rust', 'targa'
    ],
    "PM": ["pm"],
}

# def assign_label(text, dictionary):
#     text = text.lower() # Convert to lowercase for case-insensitive matching
#     for label, keywords in dictionary.items():
#         if any(word in text for word in keywords):
#             return label
#     return "Unknown" # Return 'Unknown' if no match is found

def clean_string(text):
    if isinstance(text, str):
        # Replace non-breaking spaces and then strip regular whitespace
        text= text.replace('\xa0', ' ').strip()
        text = re.sub(r'\s+', ' ', text)
        return text.strip().lower()
    return text

def assign_label(text, dictionary):
    text = clean_string(text).lower()
    print(f"Assigning label for text: {text}")
    found_labels = []
    for label, keywords in dictionary.items():
        # Create a single regex pattern for all keywords in the list
        # Using '\b' to match whole words and '|' for 'OR'
        pattern = r'\b(' + '|'.join(re.escape(word) for word in keywords) + r')\b'
        if re.search(pattern, text):
            found_labels.append(label)
    return found_labels if found_labels else ['MECH']

# Example usage with the optimized f
# Example usage
# text_to_check = "I need to replace a flat tire."
# assigned_label = assign_label(text_to_check, component_dict)
# print(f"The text is related to: {assigned_label}")

# text_to_check_2 = "The engine is making a strange noise."
# assigned_label_2 = assign_label(text_to_check_2, component_dict)
# print(f"The text is related to: {assigned_label_2}")


def segregate_by_type(data):
    """
    Segregates the input data by the type of maintenance.

    Args:
        data (list): The input data to segregate.

    Returns:
        dict: A dictionary with maintenance types as keys and lists of corresponding rows as values.
    """
    
    segregated = {}
    for row in data:
        maintenance_type = row[7]
        assigned_labels = assign_label(maintenance_type, component_dict)
        label_string = ', '.join(assigned_labels)
        if label_string not in segregated:
            segregated[label_string] = []
        segregated[label_string].append(row)
    pprint.pprint(segregated)
    
    return segregated

def create_formatted_excel(data, new , title, output_filename="formatted_report.xlsx"):
    """
    Creates an Excel file from a nested dictionary, formatting the 'total' rows in red
    and merging date cells.
    """
    # pprint.pprint(data)
    type_summary = {"NMC COUNT AS PER PROJECT REPORT":0, "REPORTED TO W/SHOP":0,"REPORTED AFTER 10 AM ON THE SAME DAY":0,
                "RELEASED":0, "PENDING":0, "MECH":{"in":0, "out":0, "pending":0},
                "ELE":{"in":0, "out":0, "pending":0}, "BODY":{"in":0, "out":0, "pending":0},
                "TYRE":{"in":0, "out":0, "pending":0},"PM":{"in":0, "out":0, "pending":0}}

    if new==1:
        wb=load_workbook(output_filename)
        ws=wb.create_sheet(title, 1)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title
        
    # Handle the case where the file is not found
   
        
   
    
    # 1. Get all unique statuses (including 'cost') for headers
    all_statuses = []
    all_statuses.extend([
        "Type","S.No","Date of B/D", "Time of B/D", "Driver Emp No #", "Asset Number", "Bus No.", "Location",
        "Type Of B/D", "Operation Action", "Re-Operations time", "W/S Comments", "workshop reported time"
    ])

    # sorted_statuses = sorted(list(all_statuses))
    # bring_to_last=[sorted_statuses.pop(sorted_statuses.index("Workshop Checked In")),
    # sorted_statuses.pop(sorted_statuses.index("Total WO")),
    # sorted_statuses.pop(sorted_statuses.index("Material cost")),]
    # sorted_statuses= [bring_to_last[1]]+sorted_statuses+[bring_to_last[0]]+[bring_to_last[2]]
    

    # Define the merged cell range (A1 to Q1 for 17 columns)
    merge_range = 'A1:L1'

    # Merge the cells
    ws.merge_cells(merge_range)

    # Set the value of the merged cell
    ws['A1'] = "Work Order Status Report"

    # Create a grey fill pattern and a white font
    grey_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    white_font = Font(color="FFFFFF")

    # Apply the fill and font to the merged cell
    ws['A1'].fill = grey_fill
    ws['A1'].font = white_font

    # Center the text in the merged cell
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')


    # 2. Define and style the header row
    # header = ['Date', 'Project'] + sorted_statuses
    ws.append(list(all_statuses))

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')
    for cell in ws[2]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border=bottom_right_border

    # 3. Write data rows, applying red font to 'total' rows
    red_font = Font(color="FF0000")  # Red font color
    
    # sorted_dates = sorted(data.keys())
    row_count = 3  # Start row for data is 2
    # print(data)
    for row in data:
        start_row = row_count
        
        for val in data[row]:
            copy=list(val)
            type_summary["NMC COUNT AS PER PROJECT REPORT"]+=1
            
            if copy[11]!="Not reported":
                type_summary["REPORTED TO W/SHOP"]+=1
                if "MECH" in row:
                    type_summary["MECH"]["in"]+=1
                    if copy[-1]==0:
                        type_summary["MECH"]["pending"]+=1
                        type_summary["PENDING"]+=1
                    elif copy[-1]==1:
                        type_summary["MECH"]["out"]+=1
                        type_summary["RELEASED"]+=1
                if  "ELE" in row:
                    type_summary["ELE"]["in"]+=1
                    if copy[-1]==0:
                        type_summary["ELE"]["pending"]+=1
                        type_summary["PENDING"]+=1
                    elif copy[-1]==1:
                        type_summary["ELE"]["out"]+=1
                        type_summary["RELEASED"]+=1

                if "BODY" in row:
                    type_summary["BODY"]["in"]+=1
                    if copy[-1]==0:
                        type_summary["BODY"]["pending"]+=1
                        type_summary["PENDING"]+=1
                    elif copy[-1]==1:
                        type_summary["BODY"]["out"]+=1
                        type_summary["RELEASED"]+=1
                if "TYRE" in row:
                    type_summary["TYRE"]["in"]+=1
                    if copy[-1]==0:
                        type_summary["TYRE"]["pending"]+=1
                        type_summary["PENDING"]+=1
                    elif copy[-1]==1:
                        type_summary["TYRE"]["out"]+=1
                        type_summary["RELEASED"]+=1


           
            copy.pop(-1)
            val=tuple(copy)
            ws.append((row,)+val)
            last_row_num = ws.max_row

            # # Apply the fill to each cell in the row
            for cell in ws[last_row_num]:
               cell.border =right_left_border
               cell.alignment=Alignment(horizontal='center', vertical='center')
                
            

            row_count += 1
        
        # Merge cells for the 'Date' column after all projects for a date are added
        ws.append([])
        row_count += 1
        print(last_row_num==ws.max_row)
        last_row_num = ws.max_row+1
        for cell in ws[last_row_num]:
               cell.border =bottom_top_border
               cell.alignment=Alignment(horizontal='center', vertical='center')
        
        end_row = row_count - 1
        if start_row < end_row:
            ws.merge_cells(start_row=start_row, end_row=end_row, start_column=1, end_column=1)
            # Center the merged cell text
            merged_cell = ws.cell(row=start_row, column=1)
            # merged_cell.fill=colors[count%4]

            ws.cell(row=start_row, column=1).alignment = Alignment(vertical='center')

    # 4. Adjust column widths for readability
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width
    

    for cell in ws['A']:
    # Apply the border to each cell
        cell.border = bottom_top_right_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions["A"].width=10

    # for cell in ws['L']:
    # # Apply the border to each cell
    #     cell.border = right_border


    # 5. Save the workbook
    pprint.pprint(type_summary)
    try:
        # delete_empty_columns(ws)
        # delete_empty_rows(ws)
        wb.save(output_filename)
        print(f"Successfully created '{title}'")
        return type_summary
        
    
    except Exception as e:
        print(f"Error saving the file: {e}")



