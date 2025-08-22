import openpyxl
import openpyxl.utils
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_summary_report(total_summary_data):
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Define styles to match the image
    header_font = Font(bold=True, size=10, color="FFFFFF")  # White text
    data_font = Font(size=10)  # Regular font for data
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue background
    sub_header_fill = PatternFill(start_color="B7D7EA", end_color="B7D7EA", fill_type="solid")  # Light blue

    # Define different border styles
    thick_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )

    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_alignment = Alignment(horizontal='center', vertical='center')

    # --- Header Section (Starting from A1) ---

    # Row 1 - Main headers
    ws['A1'] = 'S.\nNO'
    ws.merge_cells('A1:A2')
    ws['B1'] = 'DATE'
    ws.merge_cells('B1:B2')
    ws['C1'] = 'NMC\nCOUNT\nAS PER\nPROJECT\nREPORT'
    ws.merge_cells('C1:C2')
    ws['D1'] = 'DELIVERY STATUS BETWEEN\n5 AM TO 10 AM'
    ws.merge_cells('D1:F1')
    ws['G1'] = 'REPORTED DEFECTS AND RELEASED STATUS TILL 10 AM'
    ws.merge_cells('G1:X1')
    ws['Y1'] = 'VARIANCE BETWEEN\nPROJECT REPORT AND WORKSHOP REPORT'
    ws.merge_cells('Y1:AC1')

    # Row 2 - Sub headers for delivery status
    ws['D2'] = 'REPORTED\nTO\nW/SHOP'
    ws['E2'] = 'RELEASED'
    ws['F2'] = 'PENDING'

    # Defect categories in row 2
    ws['G2'] = 'PM'
    ws.merge_cells('G2:I2')
    ws['J2'] = 'TYRE'
    ws.merge_cells('J2:L2')
    ws['M2'] = 'MECH'
    ws.merge_cells('M2:O2')
    ws['P2'] = 'BODY'
    ws.merge_cells('P2:R2')
    ws['S2'] = 'A C & ELE'
    ws.merge_cells('S2:U2')
    ws['V2'] = 'W/S OR\nWINDOW\nGLASS'
    ws.merge_cells('V2:X2')

    # Variance section row 2
    ws['Y2'] = 'Total\nDifferences'
    ws.merge_cells('Y2:Y3')
    ws['Z2'] = 'REPORTED\nAFTER 10 AM\nON THE\nSAME DAY'
    ws.merge_cells('Z2:Z3')
    ws['AA2'] = 'RELEASED\nON THE\nSAME DAY'
    ws.merge_cells('AA2:AA3')
    ws['AB2'] = 'PENDING'
    ws.merge_cells('AB2:AB3')
    ws['AC2'] = 'NOT\nREPORTED\nON THE\nSAME DAY'
    ws.merge_cells('AC2:AC3')

    # Row 3 - IN/OUT/PENDING headers for defect categories
    sub_headers = ['IN', 'OUT', 'PENDING']
    col_ranges = [
        (7, 9),   # PM (G-I)
        (10, 12), # TYRE (J-L) 
        (13, 15), # MECH (M-O)
        (16, 18), # BODY (P-R)
        (19, 21), # A C & ELE (S-U)
        (22, 24)  # W/S OR WINDOW GLASS (V-X)
    ]

    for start_col, end_col in col_ranges:
        for i, header in enumerate(sub_headers):
            ws.cell(row=3, column=start_col + i).value = header

    # Sample data rows (starting from row 4)
    TNMC=0
    TREPORTED=0
    TRELEASED=0
    TPENDING=0
    TMECH_IN=0
    TMECH_OUT=0
    TMECH_PENDING=0
    TMECH_TOTAL=0
    TTYRE_IN=0
    TTYRE_OUT=0
    TTYRE_PENDING=0
    TELE_IN=0
    TELE_OUT=0
    TELE_PENDING=0
    TBODY_IN=0
    TBODY_OUT=0
    TBODY_PENDING=0
    TPM_IN=0
    TPM_OUT=0
    TPM_PENDING=0
    TPM_TOTAL=0
    Tdiff=0
    Tafter_10=0
    


    data_rows = [
        # [1, 21, '7-Aug', 18, 18, 0, 0, 0, 0, 2, 2, 0, 14, 14, 0, 0, 0, 0, 2, 2, 0, 0, 0, 0, 3, 0, 0, 0, 3],
        # [2, 21, '9-Aug', 21, 21, 0, 0, 0, 0, 3, 3, 0, 13, 13, 0, 1, 1, 0, 4, 4, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        # [3, 21, '10-Aug', 17, 15, 1, 0, 0, 0, 3, 3, 0, 6, 6, 0, 3, 2, 1, 5, 5, 0, 0, 0, 0, 4, 0, 0, 0, 4],
        # [4, 21, '11-Aug', 20, 19, 1, 0, 0, 0, 4, 4, 0, 9, 9, 0, 0, 0, 0, 7, 6, 1, 0, 0, 0, 1, 0, 0, 0, 1],
        # [5, 26, '12-Aug', 21, 20, 1, 0, 0, 0, 2, 2, 0, 17, 16, 1, 0, 0, 0, 2, 2, 0, 0, 0, 0, 5, 0, 0, 0, 5],
        # [6, 37, '13-Aug', 20, 20, 0, 0, 0, 0, 3, 3, 0, 12, 12, 0, 0, 0, 0, 5, 5, 0, 0, 0, 0, 17, 0, 0, 0, 17]
    ]


    for idex, datee in enumerate(total_summary_data, start=1):
        NMC=total_summary_data[datee]["NMC COUNT AS PER PROJECT REPORT"]
        REPORTED=total_summary_data[datee].get("REPORTED TO W/SHOP", 0)
        RELEASED=total_summary_data[datee].get("RELEASED", 0)
        PENDING=total_summary_data[datee].get("PENDING", 0)

        MECH_IN=total_summary_data[datee]["MECH"].get("in", 0)
        MECH_OUT=total_summary_data[datee]["MECH"].get("out", 0)
        MECH_PENDING=total_summary_data[datee]["MECH"].get("pending", 0)

        ELE_IN=total_summary_data[datee]["ELE"].get("in", 0)
        ELE_OUT=total_summary_data[datee]["ELE"].get("out", 0)
        ELE_PENDING=total_summary_data[datee]["ELE"].get("pending", 0)

        BODY_IN=total_summary_data[datee]["BODY"].get("in", 0)
        BODY_OUT=total_summary_data[datee]["BODY"].get("out", 0)
        BODY_PENDING=total_summary_data[datee]["BODY"].get("pending", 0)

        TYRE_IN=total_summary_data[datee]["TYRE"].get("in", 0)
        TYRE_OUT=total_summary_data[datee]["TYRE"].get("out", 0)
        TYRE_PENDING=total_summary_data[datee]["TYRE"].get("pending", 0)

        PM_IN=total_summary_data[datee]["PM"].get("in", 0)
        PM_OUT=total_summary_data[datee]["PM"].get("out", 0)
        PM_PENDING=total_summary_data[datee]["PM"].get("pending",0)

        diff=NMC-REPORTED
        after_10=total_summary_data[datee]["REPORTED AFTER 10 AM ON THE SAME DAY"]

        TNMC+=NMC
        TREPORTED+=REPORTED 
        TRELEASED+=RELEASED
        TPENDING+=PENDING

        TMECH_IN+=MECH_IN
        TMECH_OUT+=MECH_OUT
        TMECH_PENDING+=MECH_PENDING

        TTYRE_IN+=TYRE_IN
        TTYRE_OUT+=TYRE_OUT
        TTYRE_PENDING+=TYRE_PENDING

        TELE_IN+=ELE_IN
        TELE_OUT+=ELE_OUT
        TELE_PENDING+=ELE_PENDING

        TBODY_IN+=BODY_IN
        TBODY_OUT+=BODY_OUT
        TBODY_PENDING+=BODY_PENDING

        TPM_IN+=PM_IN
        TPM_OUT+=PM_OUT
        TPM_PENDING+=PM_PENDING

        TPM_TOTAL+=PM_IN+PM_OUT+PM_PENDING
        Tdiff+=diff
        Tafter_10+=after_10


        row=[idex, datee, NMC, REPORTED, RELEASED, PENDING,
              PM_IN, PM_OUT, PM_PENDING,
              TYRE_IN, TYRE_OUT, TYRE_PENDING,
              MECH_IN, MECH_OUT, MECH_PENDING,
              BODY_IN, BODY_OUT, BODY_PENDING,
              ELE_IN, ELE_OUT, ELE_PENDING,
              0, 0, 0,  # Placeholder for W/S OR WINDOW GLASS
              diff, after_10, RELEASED, PENDING,0]
        data_rows.append(row)


  
    # Add data rows starting from row 4
    for row_idx, row_data in enumerate(data_rows, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    # Add TOTAL row
    total_row = len(data_rows) + 4
    ws.cell(row=total_row, column=1).value = 'TOTAL'
    total_data = [TNMC, TREPORTED, TRELEASED, TPENDING,
                  TPM_IN, TPM_OUT, TPM_PENDING, TPM_TOTAL,
                  TMECH_IN, TMECH_OUT, TMECH_PENDING,
                  TTYRE_IN, TTYRE_OUT, TTYRE_PENDING,
                  TELE_IN, TELE_OUT, TELE_PENDING,
                  TBODY_IN, TBODY_OUT, TBODY_PENDING,
                  0, 0, 0,  # Placeholder for W/S OR WINDOW GLASS
                  Tdiff, Tafter_10, TRELEASED, TPENDING, ]
    for col_idx, value in enumerate(total_data, start=2):
        ws.cell(row=total_row, column=col_idx).value = value

    # Apply formatting to all header cells (rows 1-3) including merged cell ranges
    for row in range(1, 4):
        for col in range(1, 30):  # Adjust range as needed
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = header_border

    # Define merged cell ranges and apply proper borders
    merged_ranges = [
        # Row 1 merged cells
        ('A1:A2', 'thick'),  # S.NO
        ('B1:B2', 'thick'),  # DATE  
        ('C1:C2', 'thick'),  # NMC COUNT
        ('D1:F1', 'thick'),  # DELIVERY STATUS
        ('G1:X1', 'thick'),  # REPORTED DEFECTS
        ('Y1:AC1', 'thick'), # VARIANCE
        
        # Row 2 merged cells
        ('G2:I2', 'medium'), # PM
        ('J2:L2', 'medium'), # TYRE
        ('M2:O2', 'medium'), # MECH
        ('P2:R2', 'medium'), # BODY
        ('S2:U2', 'medium'), # A C & ELE
        ('V2:X2', 'medium'), # W/S OR WINDOW GLASS
        ('Y2:Y3', 'medium'), # Total Differences
        ('Z2:Z3', 'medium'), # REPORTED AFTER 10 AM
        ('AA2:AA3', 'medium'), # RELEASED ON THE SAME DAY
        ('AB2:AB3', 'medium'), # PENDING
        ('AC2:AC3', 'medium'), # NOT REPORTED ON THE SAME DAY
    ]

    # Apply borders to merged cell ranges
    for cell_range, border_style in merged_ranges:
        # Parse the range
        start_cell, end_cell = cell_range.split(':')
        
        # Extract column letters and row numbers
        start_col_letter = ''.join([c for c in start_cell if c.isalpha()])
        start_row = int(''.join([c for c in start_cell if c.isdigit()]))
        end_col_letter = ''.join([c for c in end_cell if c.isalpha()])
        end_row = int(''.join([c for c in end_cell if c.isdigit()]))
        
        # Convert column letters to numbers
        start_col = openpyxl.utils.column_index_from_string(start_col_letter)
        end_col = openpyxl.utils.column_index_from_string(end_col_letter)
        
        # Apply border to all cells in the range
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                
                # Determine which borders should be thick
                left_style = border_style if col == start_col else 'thin'
                right_style = border_style if col == end_col else 'thin'
                top_style = border_style if row == start_row else 'thin'
                bottom_style = border_style if row == end_row else 'thin'
                
                cell.border = Border(
                    left=Side(style=left_style),
                    right=Side(style=right_style),
                    top=Side(style=top_style),
                    bottom=Side(style=bottom_style)
                )

    # Apply formatting to data cells
    for row in range(4, total_row + 1):
        for col in range(1, 30):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = data_alignment
            cell.font = data_font

    # Set row heights for better appearance
    for row in range(1, 4):  # Header rows
        ws.row_dimensions[row].height = 45
    for row in range(4, 11):  # Data rows
        ws.row_dimensions[row].height = 20

    # Add summary section at bottom
    summary_row = total_row + 2
    ws.merge_cells('B{}:C{}'.format(summary_row, summary_row))
    ws.merge_cells('F{}:G{}'.format(summary_row, summary_row))
    ws.merge_cells('J{}:K{}'.format(summary_row, summary_row))
    ws.merge_cells('N{}:P{}'.format(summary_row, summary_row))
    



 
    ws.cell(row=summary_row, column=2).value = 'NO: OF BUSES\nREPORTED WITHIN 5 HRS'
    ws.cell(row=summary_row, column=4).value = 117
    ws.cell(row=summary_row, column=6).value = 'NO: OF BUSES\nRELEASED'
    ws.cell(row=summary_row, column=8).value = 114
    ws.cell(row=summary_row, column=10).value = 'OUTPUT\nEFFICIENCY'
    ws.cell(row=summary_row, column=12).value = '97%'
    ws.cell(row=summary_row, column=14).value = 'NO: OF BUSES NOT\nREPORTED ON THE SAME DAY'
    ws.cell(row=summary_row, column=17).value = 29

    # Apply formatting to summary section including merged cell borders
    summary_merged_ranges = [
        ('B{}:C{}'.format(summary_row, summary_row)),
        ('F{}:G{}'.format(summary_row, summary_row)),
        ('J{}:K{}'.format(summary_row, summary_row)),
        ('N{}:P{}'.format(summary_row, summary_row))
    ]
    
    # Apply borders to summary merged ranges
    for cell_range in summary_merged_ranges:
        start_cell, end_cell = cell_range.split(':')
        
        # Extract column letters and row numbers
        start_col_letter = ''.join([c for c in start_cell if c.isalpha()])
        start_row = int(''.join([c for c in start_cell if c.isdigit()]))
        end_col_letter = ''.join([c for c in end_cell if c.isalpha()])
        end_row = int(''.join([c for c in end_cell if c.isdigit()]))
        
        # Convert column letters to numbers
        start_col = openpyxl.utils.column_index_from_string(start_col_letter)
        end_col = openpyxl.utils.column_index_from_string(end_col_letter)
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thick_border
    
    for col in [2, 4, 6, 8, 10, 12, 14, 17]:
        cell = ws.cell(row=summary_row, column=col)
        cell.border = thick_border
        cell.alignment = center_alignment
        if col in [2, 6, 10, 14]:  # Labels
            cell.fill = sub_header_fill
            cell.font = Font(bold=True, size=9)
        else:  # Values
            cell.font = Font(bold=True, size=10)

    # Set height for summary row
    ws.row_dimensions[summary_row].height = 50

    # Auto-adjust column widths to make it more readable
    column_widths = {
        1: 4,   # S.NO
        2: 14,   # NMC COUNT
        3: 8,   # DATE
        4: 10,  # REPORTED TO W/SHOP
        5: 8,   # RELEASED
        6: 8,   # PENDING
    }

    # Set widths for defect category columns (PM, TYRE, MECH, BODY, A C & ELE, W/S)
    for col in range(7, 25):  # Columns G through X
        column_widths[col] = 5

    # Set widths for variance columns
    for col in range(25, 30):  # Columns Y through AC
        column_widths[col] = 8

    # Apply column widths
    for col, width in column_widths.items():
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = width

    # Save the workbook
    wb.save('status_project_report.xlsx')
print("✅ Status Project Report table created successfully!")
print("📁 File saved as: status_project_report.xlsx")
