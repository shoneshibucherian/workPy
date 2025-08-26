import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pprint
import os
def create_daily_summary_table(data, newF=1):
    # Create a new workbook and select the active worksheet
    # pprint.pprint(data)


    #find the path of the output directory
    output_dir = os.path.join(os.getcwd(), 'output')
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)



    if newF==1:
        wb=openpyxl.load_workbook(os.path.join(output_dir, "Daily_Breakdown_Report.xlsx"))
        ws=wb.create_sheet("Summary",0)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"

    # Define styles
    header_font = Font(bold=True, size=11, color="000000")  # Black text
    data_font = Font(size=10)
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray background
    
    # Define border styles
    thick_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # --- Create Headers ---
    
    # Row 1 - Main headers
    headers = [
        "DATE",
        "TIMING", 
        "NMC COUNT AS PER\nPROJECT REPORT",
        "REPORTED TO\nWORKSHOP",
        "B/D ATTENDED AT\nLOCATION",
        "NOT REPORTED ON\nTHE SAME DAY"
    ]
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thick_border
    
    # Sample time periods for demonstration
    # time_periods = [
    #     "From 5 AM to 6 AM",
    #     "From 6 AM TO 7 AM", 
    #     "From 7 AM TO 8 AM",
    #     "From 8 AM TO 9 AM",
    #     "From 9 AM TO 10 AM"
    # ]
    time_periods = {f"FROM {hour} AM TO {hour+1} AM":{"nmc":0,"reported":0,"bd_attended":0,"not_reported":0} for hour in range(5, 10)}
    
    # Sample dates
    dates = [date for date in data.keys()]
    
    current_row = 2
    
    # Add sample data structure
    for date in dates:
        # First row for this date
        first_time_period = True
        
        for time_period in time_periods:
            # Write date only in the first row for each date
            if first_time_period:
                # Merge cells for date column to span all time periods for this date
                start_row = current_row
                end_row = current_row + len(time_periods) - 1
                ws.merge_cells(f'A{start_row}:A{end_row}')
                
                date_cell = ws.cell(row=start_row, column=1, value=date)
                date_cell.font = data_font
                date_cell.alignment = center_alignment
                date_cell.border = thin_border
                
                first_time_period = False
            
            # Write timing
            timing_cell = ws.cell(row=current_row, column=2, value=time_period)
            timing_cell.font = data_font
            timing_cell.alignment = center_alignment
            timing_cell.border = thin_border
            
            # Add placeholder data (0 for now, you can replace with actual data later)
            col_list = ["nmc","reported","bd_attended","not_reported"]
            # col_list[1]=col_list[1]-col_list[2]
            data[date][time_period]["reported"] = data[date][time_period]["reported"]-data[date][time_period]["bd_attended"]
            for col in range(3, 7):  # Columns C through F
                # print("Filling data for", date, time_period, col_list[col-3], data.get(date, {}).get(time_period, {}).get(col_list[col-3], 0))
                data_cell = ws.cell(row=current_row, column=col, value=data.get(date, {}).get(time_period, {}).get(col_list[col-3], 0))
                data_cell.font = data_font
                data_cell.alignment = center_alignment
                data_cell.border = thin_border
            
            current_row += 1
    
    # Set column widths for better appearance
    column_widths = {
        1: 12,  # DATE
        2: 20,  # TIMING
        3: 15,  # NMC COUNT
        4: 15,  # REPORTED TO WORKSHOP
        5: 18,  # B/D ATTENDED AT LOCATION
        6: 18   # NOT REPORTED ON THE SAME DAY
    }
    
    for col, width in column_widths.items():
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = width
    
    # Set row height for header
    ws.row_dimensions[1].height = 45
    
    # Set row heights for data rows
    for row in range(2, current_row):
        ws.row_dimensions[row].height = 25
    
    # Save the workbook
    filename =os.path.join(output_dir, "Daily_Breakdown_Report.xlsx")
    wb.save(filename)
    print(f"✅ Daily Summary table template created successfully!")
    print(f"📁 File saved as: {filename}")
    
    return filename

def add_data_to_daily_summary(data_dict):
    """
    Function to add actual data to the daily summary table
    
    data_dict format:
    {
        "14/8/2025": {
            "From 5 AM to 6 AM": {
                "nmc_count": 8,
                "reported_to_workshop": 7,
                "bd_attended": 0,
                "not_reported": 1
            },
            # ... more time periods
        },
        # ... more dates
    }
    """
    # This function can be implemented later when you have the actual data
    # For now, it's just a placeholder structure
    pass

# # Example usage
# if __name__ == "__main__":
#     create_daily_summary_table()
