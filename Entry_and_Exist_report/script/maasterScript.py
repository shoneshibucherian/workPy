import os
from SummaryData import insert_batch_headers
from E_Ereport import process_excel_data
from summery import create_summary_report
from dailySummary import create_daily_summary_table




col_names=()
    #read the data in column E store it in set to ensure unique values

# maintenance_schedule = process_excel_data(file_path)

output_dir = os.path.join(os.getcwd(), 'output')
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

    
    
    
 

if os.path.exists(os.path.join(output_dir, "Entry_and_Exit_of_Bus.xlsx")):
    os.remove(os.path.join(output_dir, "Entry_and_Exit_of_Bus.xlsx"))
if os.path.exists(os.path.join(output_dir, "Daily_Breakdown_Report.xlsx")):
    os.remove(os.path.join(output_dir, "Daily_Breakdown_Report.xlsx"))

total_summary = {}
hour_summary = {}
daily_status_dir= os.path.join(os.getcwd(), 'daily_status')
file_name = 'input.xlsx'
file_path = os.path.join(daily_status_dir, file_name)

if not os.path.exists(daily_status_dir):
    print("not exists")


file_list=[]
for file in os.listdir(daily_status_dir):
    if file.endswith('10AM.xlsx'):
        file_list.append(os.path.join(daily_status_dir, file))
    else:
        #anyother file, rename it to input.xlsx
        if (os.path.exists(os.path.join(daily_status_dir, file_name)))==False:
            file_path=os.rename(os.path.join(daily_status_dir, file), os.path.join(daily_status_dir, file_name))
        if (os.path.exists(os.path.join(daily_status_dir, file_name)))==False:
            print("File not found:", file_name)
            #how to exit the program
            exit(1)
        # file_list.append(os.path.join(daily_status_dir, 'input.xlsx'))

print("Files in AQAQ daily_status directory:")  
for file in file_list:
    print(file)
    sheet_name=file[file.index("daily_status")+len("daily_status")+1:file.index("Daily")]
    # read the unique values in column E in the excel file
    from openpyxl import load_workbook
    input_wbook = load_workbook(file)
    inpute_ws = input_wbook.active
    col_names = set()
    Sdate=inpute_ws['B2'].value.date().isoformat()
    print("Processing date:", Sdate)
    for row in inpute_ws.iter_rows(min_row=2, max_col=5, values_only=True):
        #before adding remove all spaces in the value
        if row[4]:
            col_names.add(row[4].replace(" ", ""))
    print("Unique values in column E:", col_names)
    maintenance_schedule, countert = process_excel_data(col_names,file_path,Sdate)
    Batches_hour,summary_per_day, date=insert_batch_headers(countert,maintenance_schedule,file, os.path.join(output_dir, "Daily_Breakdown_Report.xlsx"), sheet_name,0)
    total_summary[date] = summary_per_day
    hour_summary[date] = Batches_hour

print("Total Summary:")
for date, summary in total_summary.items(): 
    # print(f"Date: {date.date()}")
    for key, value in summary.items():
        print(f"{key}: {value}")


create_summary_report(total_summary)
create_daily_summary_table(hour_summary)
import time
time.sleep(3)