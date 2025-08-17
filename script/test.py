from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from datetime import datetime

def get_hour_range(time_str):
    """Extract hour from time string like '5:25'."""
    try:
        print(time_str)
        print(time_str.index(":"))
        dt = int(time_str[:time_str.index(":")])
        print(dt)
        return dt
    except ValueError:
        return None

def insert_batch_headers(filename, output_filename):
    wb = load_workbook(filename)
    ws = wb.active

    # Collect data rows (starting from row 2 assuming headers in row 1)
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header = [cell.value for cell in ws[1]]
    time_col_idx = 3  # 'Time of B/D' is column B (index starts from 1)
    num_cols = len(header)

    # Group rows by hour
    batches = {}
    for row in rows:
        time_val = row[time_col_idx - 1]
        print(time_val)
        print(type(time_val))
        hour = get_hour_range(str(time_val))
        print(hour)
        if hour is not None:
            if hour not in batches:
                batches[hour] = []
            batches[hour].append(row)

    # Sort the batches by hour
    sorted_hours = sorted(batches.keys())

    # Clear the sheet and rewrite with batch headers
    ws.delete_rows(2, ws.max_row)
    current_row = 2

    for idx, hour in enumerate(sorted_hours):
        next_hour = hour + 1
        label = f"{idx+1}ST BATCH - FROM {hour} AM TO {next_hour} AM" if idx == 0 else \
                f"{idx+1}ND BATCH - FROM {hour} AM TO {next_hour} AM" if idx == 1 else \
                f"{idx+1}RD BATCH - FROM {hour} AM TO {next_hour} AM" if idx == 2 else \
                f"{idx+1}TH BATCH - FROM {hour} AM TO {next_hour} AM"

        # Insert merged header row
        ws.insert_rows(current_row)
        merge_range = f"A{current_row}:{get_column_letter(num_cols)}{current_row}"
        ws.merge_cells(merge_range)
        cell = ws.cell(row=current_row, column=1, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        current_row += 1

        # Insert batch rows
        for row_data in batches[hour]:
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=current_row, column=col_idx, value=value)
            current_row += 1

    wb.save(output_filename)
    print(f"✅ Done! Modified file saved as: {output_filename}")

# ======== USAGE ===========
input_file = "7 August Daily Breakdown Report 5-10AM.xlsx"       # Replace with your filename
output_file = "output_batched.xlsx"

insert_batch_headers(input_file, output_file)
