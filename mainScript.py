from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment,PatternFill,Border,Side
from datetime import datetime
import openpyxl
import pprint
dict={}

date_project={}

light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
light_purple_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
light_pink_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
colors=[light_blue_fill,light_green_fill,light_purple_fill, light_pink_fill]
bottom_border = Border(bottom=Side(style='thick'))
left_border = Border(left=Side(style='thick'))
right_border = Border(right=Side(style='thin'))
bottom_right_border = Border(right=Side(style='thick'), 
                     bottom=Side(style='thick'))
right_left_border=Border(right=Side(style='thick'), 
                     left=Side(style='thick'))
bottom_top_right_border=Border(right=Side(style='thick'), 
                     bottom=Side(style='thick'),
                     top=Side(style="thick"))


count=0



proj={"TC-TERMINAL":"TC-TERMINL", "TC-OPEN":"TC-OPEN",
 "PTS- SPOT RENTAL":"PTS-SPRENT","TC- US Government":"TC-US Govt",
 "KGL Food Services - PV Project":"FS-PV","TC-CARGO & CONTAINER":"TC-KT",
"TC-OVRLAND":"TC-OVRLAND",
 "TC-OVERLAND COMMERCIAL":"TC-OVRLAND","TC-KNPC FUEL TRANSPORTATION":"TC-KNPCFT",
 "TC-MHE CRANE LEASE":"TC-MHECRAN","TS-MAB WORKSHOP":"TS-MAB WS",
 "TC-BULK LIQUID TRANSPORT":"TC-BLKLT","TC-Energy Solutions":"TC-ES",
"ZOUD":"ZOUD","PTS-USG BUS SERVICES":"PTS-USGBUS",
"HOLDING PROCUREMENT":"H-PROCURE","FS-CATERING":"FS-CATER",
"AL-OPERATIONS ASSET POOL":"AL-ASTPOOL","PTS-PT SUPPORT":"PTS-PT SUPPORT",
"KGL Logistics - KOC - 19054114":"KGLL-KOC",
"Al MURABITOUN AL KHALIJIAH":"Al MURABITOUN AL KHALIJIAH","LOGISTICS WAREHOUSE 3PL MAB":'LOGISTICS WAREHOUSE 3PL MAB',
"MURABITOUN":"MURABITOUN","LOG-WH3PLM":"LOG-WH3PLM","KGLL-KOC":"KGLL-KOC"}




def hide_unused_cells(sheet):
    """Hides all rows and columns not containing data."""
    
    # Get the last row and column that have content
    max_row = sheet.max_row
    max_col = sheet.max_column

    # Hide all rows from the next row to the end of the sheet
    for row_num in range(max_row + 1, sheet.max_row + 10):  # A small range is fine for visual example
        sheet.row_dimensions[row_num].hidden = True
        
    # Hide all columns from the next column to the end of the sheet
    for col_num in range(max_col + 1, sheet.max_column + 10): # A small range for example
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].hidden = True



def add_dicts(dict1, dict2):
    result = dict1.copy()
    for key, value in dict2.items():
        result[key] = result.get(key, 0) + value
    
    return result


def creatDic(input):
    wb = load_workbook(input)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header = [cell.value for cell in ws[1]]
    status_col_idx = 8 # 'Time of B/D' is column B (index starts from 1)
    num_cols = len(header)

    for row in rows:
        status = row[status_col_idx - 1]
        
        date= row[3].strftime("%Y-%d-%m")
        project=row[20]
        
        
        
        if date not in dict:
            dict[date]={"Total":{"Total WO":0}}
        if project not in dict[date]:
            dict[date][project]={"Total WO":0}

        if status not in dict[date]["Total"]:
            dict[date]["Total"][status]=1
            
        else:
            dict[date]['Total'][status]+=1
        

        
        if status not in dict[date][project]:
            dict[date][project][status]=1
            
        else:
            dict[date][project][status]+=1
        
        dict[date][project]["Total WO"]+=1
        dict[date]["Total"]["Total WO"]+=1



            # status_dict={status:1}
        # dict[date]=add_dicts(dict[date],status_dict)


def create_formatted_excel(data, output_filename="formatted_report.xlsx"):
    """
    Creates an Excel file from a nested dictionary, formatting the 'total' rows in red
    and merging date cells.
    """
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Status Report"
    
    # 1. Get all unique statuses (including 'cost') for headers
    all_statuses = set()
    for date_data in data.values():
        for project_data in date_data.values():
            all_statuses.update(project_data.keys())

    sorted_statuses = sorted(list(all_statuses))
    bring_to_last=[sorted_statuses.pop(sorted_statuses.index("Workshop Checked In")),
    sorted_statuses.pop(sorted_statuses.index("Total WO"))]
    sorted_statuses= [bring_to_last[1]]+sorted_statuses+[bring_to_last[0]]
    

    # Define the merged cell range (A1 to Q1 for 17 columns)
    merge_range = 'A1:L1'

    # Merge the cells
    ws.merge_cells(merge_range)

    # Set the value of the merged cell
    ws['A1'] = "Work Order Status Report"

    # Create a grey fill pattern and a white font
    grey_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    white_font = Font(color="FFFFFF")

    # Apply the fill and font to the merged cell
    ws['A1'].fill = grey_fill
    ws['A1'].font = white_font

    # Center the text in the merged cell
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')


    # 2. Define and style the header row
    header = ['Date', 'Project'] + sorted_statuses
    ws.append(header)

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')
    for cell in ws[2]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border=bottom_right_border

    # 3. Write data rows, applying red font to 'total' rows
    red_font = Font(color="FF0000")  # Red font color
    
    sorted_dates = sorted(data.keys())
    row_count = 3  # Start row for data is 2
    for date in sorted_dates:
        global count
        count+=1
        start_row = row_count
        tmp_list=list(data[date].items())
        
        filt=[(i,j) for i, j in tmp_list if i == "Total"]
        tmp_list.pop(tmp_list.index(filt[0]))
        tmp_list.append(filt[0])
        for project, project_data in tmp_list:
            row_data = [date, project]
            for status in sorted_statuses:
                row_data.append(project_data.get(status, 0))
            ws.append(row_data)
            last_row_num = ws.max_row
            ws[last_row_num]

            # # Apply the fill to each cell in the row
            for cell in ws[last_row_num]:
               cell.border =right_left_border
                
            
            # Apply red font to the entire row if the project name is 'total'
            if project == 'Total':
                for cell in ws[ws.max_row]:
                    cell.font = red_font
                    cell.fill= yellow_fill
                    cell.border=bottom_top_right_border

            row_count += 1
        
        # Merge cells for the 'Date' column after all projects for a date are added
        end_row = row_count - 1
        if start_row < end_row:
            ws.merge_cells(start_row=start_row, end_row=end_row, start_column=1, end_column=1)
            # Center the merged cell text
            merged_cell = ws.cell(row=start_row, column=1)
            # merged_cell.fill=colors[count%4]

            ws.cell(row=start_row, column=1).alignment = Alignment(vertical='center')

    # 4. Adjust column widths for readability
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width
    

    for cell in ws['A']:
    # Apply the border to each cell
        cell.border = bottom_top_right_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions["A"].width=10

    # for cell in ws['L']:
    # # Apply the border to each cell
    #     cell.border = right_border


    # 5. Save the workbook
    try:
        # delete_empty_columns(ws)
        # delete_empty_rows(ws)
        hide_unused_cells(ws)
        wb.save(output_filename)
        print(f"Successfully created '{output_filename}'")
    except Exception as e:
        print(f"Error saving the file: {e}")





def add_est_cost(input):
    wb = load_workbook(input)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header = [cell.value for cell in ws[1]]
    cost_col_idx = 21 # 'Time of B/D' is column B (index starts from 1)
    num_cols = len(header)
    for row in rows:
        cost=None
        
        
        if row[2] and row[cost_col_idx - 1]:
            cost = float(row[cost_col_idx - 1])
            
        else:
            continue
        
        project=proj[row[8]]
        # print(row[2])    
        date= row[2].strftime("%Y-%m-%d")

        
        if date not in dict:
            dict[date]={}

        if project not in dict[date]:
            dict[date][project]={}


        if "Material cost" not in dict[date]["Total"]:
            dict[date]['Total']["Material cost"]=cost
        else:
            dict[date]['Total']["Material cost"]+=cost

        if "Material cost" not in dict[date][project]:
            dict[date][project]["Material cost"]=cost
        else:
            dict[date][project]["Material cost"]+=cost
            
   

            # status_dict={status:1}
        # dict[date]=add_dicts(dict[date],status_dict)
    
def delete_empty_rows(sheet):
    for row_num in reversed(range(1, sheet.max_row + 1)):
        is_empty = all(cell.value is None for cell in sheet[row_num])
        if is_empty:
            sheet.delete_rows(row_num, 1)
            

def delete_empty_columns(sheet):
    for col_num in reversed(range(1, sheet.max_column + 1)):
        is_empty = all(sheet.cell(row=row, column=col_num).value is None for row in range(1, sheet.max_row + 1))
        if is_empty:
            sheet.delete_cols(col_num, 1)
    

status_report="status.xlsx"
cost_report="cost.xlsx"

creatDic(status_report)
add_est_cost(cost_report)
pprint.pprint(dict)
create_formatted_excel(dict, "output/status_project_report.xlsx")