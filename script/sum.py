from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from datetime import datetime
import openpyxl
dict={}

date_project={}

def add_dicts(dict1, dict2):
    result = dict1.copy()
    for key, value in dict2.items():
        result[key] = result.get(key, 0) + value
    print(result)
    return result


def creatDic(input):
    wb = load_workbook(input)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header = [cell.value for cell in ws[1]]
    status_col_idx = 8 # 'Time of B/D' is column B (index starts from 1)
    num_cols = len(header)

    for row in rows:
        status = row[status_col_idx - 1]
        
        date= row[3].strftime("%Y-%d-%m")
        
        
        if date not in dict:
            dict[date]={}
        if status not in dict[date]:
            dict[date][status]=1
        else:
            dict[date][status]+=1

            # status_dict={status:1}
        # dict[date]=add_dicts(dict[date],status_dict)


def create_excel_from_dict(data, output_filename="report.xlsx"):
    """
    Creates an Excel file from a dictionary where keys are dates and values are
    dictionaries of statuses and counts.
    """
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Status Report"

    # Get all unique statuses to use as column headers
    all_statuses = set()
    for date_data in data.values():
        all_statuses.update(date_data.keys())
    
    # Sort statuses for consistent column order
    sorted_statuses = sorted(list(all_statuses))

    # Define the header row
    header = ['Date'] + sorted_statuses
    ws.append(header)

    # Style the header
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment

    # Write the data rows
    sorted_dates = sorted(data.keys())
    for date in sorted_dates:
        row_data = [date] 
        # print(date) # Format date to a string
        # print(row_data)
        date_dict = data[date]
        for status in sorted_statuses:
            row_data.append(date_dict.get(status, 0)) # Use .get() to handle missing statuses with a default of 0
        ws.append(row_data)

    # Adjust column widths for better readability
    for col in ws.columns:
        max_length = 0
        column = col[0].column # Get the column index
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Save the workbook
    try:
        wb.save(output_filename)
        print(f"Successfully created '{output_filename}'")
    except Exception as e:
        print(f"Error saving the file: {e}")


def add_est_cost(input):
    wb = load_workbook(input)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header = [cell.value for cell in ws[1]]
    cost_col_idx = 21 # 'Time of B/D' is column B (index starts from 1)
    num_cols = len(header)

    for row in rows:
        cost=0
        
        if row[2] and row[cost_col_idx - 1]:
            cost = int(row[cost_col_idx - 1])
            
        else:
            continue
        print(row[2])    
        date= row[2].strftime("%Y-%m-%d")

        
        if date not in dict:
            dict[date]={}
        if "cost" not in dict[date]:
            dict[date]["cost"]=cost
        else:
            dict[date]["cost"]+=cost

            # status_dict={status:1}
        # dict[date]=add_dicts(dict[date],status_dict)
    
            

    

status_report="fnd_gfm_5960802.xlsx"
cost_report="FNDWRR.xlsx"

creatDic(status_report)
add_est_cost(cost_report)
# print(dict)
create_excel_from_dict(dict, "output/status_report.xlsx")