import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime
from E_Ereport import process_excel_data
from type_segregat import segregate_by_type, create_formatted_excel

# Specify the path to your Excel file
file_path = 'input.xlsx' 

# Call the function and print the result


# Print the final dictionary for verification

complete_list= []

def get_hour_range(time_str):
    """Extract hour from time string like '5:25'."""
    try:
        # print(time_str)
        # print(time_str.index(":"))
        dt = int(time_str[:time_str.index(":")])
        # print(dt)
        return dt
    except ValueError:
        return None

def insert_batch_headers(maintenance_schedule,filename, output_filename, title="Batched Report",new=1):
    complete_list=[]
    input_wbook = load_workbook(filename)
    inpute_ws = input_wbook.active

    if new==0 and os.path.exists(output_filename):
        wb=load_workbook(output_filename)
        ws=wb.create_sheet(title, 1)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title

    # Collect data rows (starting from row 2 assuming headers in row 1)
    rows = list(inpute_ws.iter_rows(min_row=2, values_only=True))
    header = [cell.value for cell in inpute_ws[1]]
    header.append("Reported Time")
    time_col_idx = 3  # 'Time of B/D' is column B (index starts from 1)
    num_cols = len(header)
    # print(rows[1][1])
    datetime_of_file = datetime.strptime(str(rows[1][1]), '%Y-%m-%d %H:%M:%S')
    # Extract the date and time components
    date_of_file = datetime_of_file.date().isoformat()
    
        
    # Group rows by hour
    batches = {}
    for row in rows:
        time_val = row[time_col_idx - 1]
        row_copy = list(row)
        
        # print(time_val)
        # print(type(time_val))
        scheduled_datetime = datetime.strptime(str(row[1]), '%Y-%m-%d %H:%M:%S')
        

            # Extract the date and time components
        schedule_date = scheduled_datetime.date().isoformat()
        row_copy[1] = schedule_date
        
        row_copy= tuple(row_copy)
        schedule_time = scheduled_datetime.time()
        hour = get_hour_range(str(time_val))
        # print(hour)
        if hour is not None:
            if hour not in batches:
                batches[hour] = []
            if row[4] in maintenance_schedule[schedule_date]:
                batches[hour].append(row_copy + (maintenance_schedule[schedule_date][row[4]]["time"],))
                complete_list.append(row_copy + (maintenance_schedule[schedule_date][row[4]]["time"],maintenance_schedule[schedule_date][row[4]]["out_time"]))
            else:
                batches[hour].append(row_copy + ("Not reported",))
                complete_list.append(row_copy + ("Not reported",-1))

    # Sort the batches by hour
    sorted_hours = sorted(batches.keys())
    segregated_dict=segregate_by_type(complete_list)


    # Clear the sheet and rewrite with batch headers
    ws.delete_rows(2, ws.max_row)
    
    # Write the updated header row with consistent formatting
    from openpyxl.styles import PatternFill
    for col_idx, header_value in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_value)
        # Apply consistent header formatting to all columns
        cell.font = Font(bold=True, color="FFFFFF")  # White text
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue background
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column widths based on content
    for col_idx, header_value in enumerate(header, start=1):
        column_letter = get_column_letter(col_idx)
        # Set minimum width based on header text length, with some padding
        if column_letter=="A":
            header_width = 3
        elif column_letter=="E"  or column_letter=="D" or column_letter=="J" or column_letter=='K':
            header_width = 9
        elif column_letter =="H" :
            header_width = 39
        else:
            header_width = len(str(header_value)) + 2
        ws.column_dimensions[column_letter].width = max(header_width, 2)  # Minimum width of 12
    
    current_row = 2

    for idx, hour in enumerate(sorted_hours):
        next_hour = hour + 1
        label = f"{idx+1}ST BATCH - FROM {hour} AM TO {next_hour} AM" if idx == 0 else \
                f"{idx+1}ND BATCH - FROM {hour} AM TO {next_hour} AM" if idx == 1 else \
                f"{idx+1}RD BATCH - FROM {hour} AM TO {next_hour} AM" if idx == 2 else \
                f"{idx+1}TH BATCH - FROM {hour} AM TO {next_hour} AM"

        # Insert merged header row
        ws.insert_rows(current_row)
        merge_range = f"A{current_row}:{get_column_letter(num_cols)}{current_row}"
        ws.merge_cells(merge_range)
        cell = ws.cell(row=current_row, column=1, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        current_row += 1

        # Insert batch rows
        for row_data in batches[hour]:
            for col_idx, value in enumerate(row_data, start=1):
                if value =="Not reported":
                    ws.cell(row=current_row, column=col_idx, value=value).font = Font(bold=True, color="FF0000") 
                else:
                    ws.cell(row=current_row, column=col_idx, value=value)
            current_row += 1

    wb.save(output_filename)
    print(f"✅ Done! Modified file saved as: {output_filename}")
    dict = create_formatted_excel(segregated_dict,1,"Segregated type Report",output_filename)
    print("Segregated type Report created",dict)
    after_10=maintenance_schedule[date_of_file]["REPORTED AFTER 10 AM ON THE SAME DAY"]
    print("REPORTED AFTER 10 AM ON THE SAME DAY",after_10)
    dict["REPORTED AFTER 10 AM ON THE SAME DAY"] = after_10
    return dict, datetime_of_file.date().isoformat()

# ======== USAGE ===========
# input_file = "14 August Daily Breakdown Report 5-10AM.xlsx"       # Replace with your filename
# output_file = "output_batched.xlsx"

# insert_batch_headers(input_file, output_file)  
