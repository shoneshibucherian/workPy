import openpyxl

# Load the workbook
wb = openpyxl.load_workbook("fnd_gfm_5960802.xlsx")

# Select the active sheet
sheet = wb.active



output= openpyxl.Workbook()

# Select the active sheet
sheet1 = output.active


# Read and print the data
for row in sheet.iter_rows(min_row=1, max_row=23, values_only=True):
    sheet1.append(row)

output.save("output.xlsx")